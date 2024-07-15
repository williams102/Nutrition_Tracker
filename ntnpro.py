"""_summary_
ntnpro.py
Author: Ben Williams

Date Modified: 14 July 2024

Version: 1.0

Summary: This module serves as a organization space for the functions and classes used in the Nutrition.py
program. The functions and classes are designed to only be used in the Nutrition.py program.

Functions:
- ask_user(prompt, kind): Forces user to input data of a specified data type.
- user_response(prompt, criteria): Forces user to match specified answers.
- search_data(data, inputdf, ): Fearches for data inside input dataframe given the category of data.
- conv_height(inches): Converts given height in inches to a feet and inches.
- add_s(text): Adds an s to the end of a string. 
- list_ave(input_list): Calculates the average of a given list.
- split_serv(text): Breaks up serving size into the numerical servings and measurement size.
- timepass(period): Calculates a starting and ending date given a time period. 
- date_index(idnum, times, datadf, meal): Creates a copy of an input dataframe with the dates as the indices. 
- combine(oldser): Combines the values of a series with duplicate indices into one with no dupilcates. 
- combdf(olddf): Combines the values of a dataframe with duplicate indices into one with no duplicates. 
- plotti(parm): Determines the title for a line plot based on specified ingredient. 
- color(parameter, ave, rec): Calculates if the intake values of a nutrition are higher than recommended and assigns a color. 
- mp_title(meal, period): Determines the title for a line plot based on the specified meal and time period. 
- ingredients_list(flist, fooddf): Breaks up string of foods and ingredients into a list. 
- daily_value(eer): Calculates the recommended daily values for nutritions based on EER. 
- percentdv(food, dvals): Calculates the percent of the daily values of a food. 

Classes:
- Profile: creates instance of a profile for a user.
- Food: creates instance of a food consumed by user.
- Day(Food): creates instance of a day for user.

Arguments:
none

Returns:
none
    _type_: _description_
"""

import pandas as pd
import numpy as np
import csv
from openpyxl import load_workbook
import math
import datetime
from datetime import datetime as dt
from dateutil import parser 
from colorama import Fore, Style

def ask_user(prompt, kind):
    while True:
        ans = input(prompt)
        try:
            ans = kind(ans)
            type(ans)
            break
        except ValueError:
            print(Style.NORMAL+Fore.RED+"\nInvalid input. Please try again.")
            print(Fore.BLACK)

    return ans

def user_response(prompt, criteria):
    answer = input(prompt)

    if answer in ['all', 'none']:
        pass
    
    elif isinstance(answer, str):
        answer = answer.strip()
        
        if answer.isalpha():
            answer = answer.lower()
        elif answer.isnumeric():
            answer = int(answer)
        elif answer.find('.') > 1:
            answer = float(answer)

        if isinstance(criteria, list):
            while answer not in criteria:
                answer = input()
                answer = answer.strip()
                
                if answer.isalpha():
                    answer = answer.lower()
                elif answer.isnumeric():
                    answer = int(answer)
                elif answer.find('.') > 1:
                    answer = float(answer)

        elif isinstance(criteria, int | float | np.int64 | np.float64):
            while answer != criteria:
                answer = input()
                answer = answer.strip()
                
                if answer.isalpha():
                    answer = answer.lower()
                elif answer.isnumeric():
                    answer = int(answer)
                elif answer.find('.') > 1:
                    answer = float(answer)
        else:
            while answer not in criteria:
                answer = input()
                answer = answer.strip()

                if answer.isalpha():
                    answer = answer.lower()
    
    return answer

def search_data(data, inputdf, category=str, disp=bool):
    ind = list()
    
    if 'sex' in inputdf.columns:
        inputdf = inputdf.copy()
        inputdf.drop_duplicates(subset = ['id'], keep = 'last', inplace = True, ignore_index = True)

    if category == 'birthday':
        data = parser.parse(data)
        data = datetime.datetime.strftime(data, '%#m/%#d/%Y')
        inputdf = inputdf.copy()

    if isinstance(data, str):
        if category.isalpha() and category.lower() == 'all':
            ind = list(range(0, len(inputdf['name'])))
            finaldf = inputdf

        elif data.isalpha() and data.lower() == 'all':
            ind = list(range(0, len(inputdf[category])))
            finaldf = inputdf

        else:
        
            for n in range(0, len(inputdf[category])):
                val = inputdf.loc[n, category]

                if data.isnumeric():
                    if int(data) == val:
                        ind.append(n)
                
                elif data.find('.') > 0:
                    if float(data) == val:
                        ind.append(n)
                    
                elif data.isalpha():
                    if isinstance(val, str):
                        if data.lower() in val.lower():
                            ind.append(n)
                else:
                    if data == val:
                        ind.append(n)
            
            finaldf = inputdf.loc[ind]

    elif isinstance(data, int):
        for n in range(0, len(inputdf[category])):
            if data == inputdf.loc[n, category]:
                ind.append(n)

    if len(ind) == 1:
        ind = ind[0]
        finaldf = inputdf.loc[ind]
    elif len(ind) == 0:
        print(Fore.RED+'\n0 results found. Please try again.')
        return False

    if disp:
        cols = inputdf.columns.tolist()     

        if isinstance(ind, list):
            print(Style.BRIGHT+f"{len(ind)} results found:\n")

            for row in range(0, len(ind)):
                print(Style.NORMAL+str(row+1)+".\t"+cols[0].title()+': '+str(finaldf.iloc[row, 0])+"\n")
                print("\t"+cols[1].title()+': '+str(finaldf.iloc[row, 1])+"\n")
                print("\t"+cols[2].title()+': '+str(finaldf.iloc[row, 2])+"\n")
            
        elif isinstance(ind, int):
            print(Style.BRIGHT+"1 result found:\n")
            print(Style.NORMAL+"1.\t"+cols[0].title()+': '+str(finaldf[cols[0]])+"\n")
            print("\t"+cols[1].title()+': '+str(finaldf[cols[1]])+"\n")
            print("\t"+cols[2].title()+': '+str(finaldf[cols[2]])+"\n")
            
    return finaldf
        
def conv_height(inches):
    feet = 0
    while True:
        inches = inches - 12
        feet = feet + 1
        if inches < 12:
            break
    
    return (feet, inches)

def add_s(text):
    if text[-1] != 's':
        text += 's'
    elif text.endswith('ss'):
        text += 'es'
    return text

def list_ave(input_list):
    float_list = [float(x) for x in input_list]
    total = sum(float_list)
    count = len(input_list)
    return total / count

def split_serv(text):
    if isinstance(text, str):
        if text.isalpha():
            return [float('NaN'), text]
            
        elif text.isnumeric():
            if '/' in text:
                frac = text.split('/')
                sz = int(frac[0]) / int(frac[1])
            elif '.' in text:
                sz = float(text)
            else: 
                sz = int(text)
            return [sz, float('NaN')]

        else:
            servings = text.split(' ')
            
            if '/' in servings[0]:
                frac = servings[0].split('/')
                sz = int(frac[0]) / int(frac[1])

                if len(servings) > 1:
                    return [sz, servings[1]]
                else:
                    return [sz, float('NaN')]
            
            elif '.' in servings[0]:
                sz = float(servings[0])

                if len(servings) > 1:
                    return [sz, servings[1]]
                else:
                    return [sz, float('NaN')]
            
            else:
                if isinstance(servings[0], int):
                    servings = [int(servings[0]), servings[1]]
                else: 
                    servings = [float(servings[0]), servings[1]]
                return servings
    elif isinstance(text, int | float | np.int64 | np.float64):
        return [text, float('NaN')]
        
def timepass(period=str):

    if period in ['custom', 'cust', 'c']:
        tstart = input(Style.NORMAL+"\nPlease enter the starting date: ")
        tstart = parser.parse(tstart)
        tstart = tstart.date()

        tend = input(Style.NORMAL+"\nPlease enter the ending date: ")
        tend = parser.parse(tend)
        tend = tend.date()

    elif period == 'today':
        tstart = datetime.date.today()
        tend = tstart

    else:
        period = split_serv(period)

        if period[1].isalpha():
            if period[1].lower() in ['days', 'day', 'd']:
                instances = period[0]
            elif period[1].lower() in ['weeks', 'week', 'wks', 'wk', 'w']:
                instances = period[0]*7
            elif period[1].lower() in ['months', 'month', 'm']:
                instances = (period[0] / 12)*365
            elif period[1].lower() in ['years', 'year', 'yrs', 'yr', 'y']:
                instances = period[0]*365
        
        tend = datetime.date.today()

        delta = datetime.timedelta(days=instances)
        tstart = tend - delta

    return tstart, tend

def date_index(idnum, times, datadf, meal):
    ddata = datadf.copy()
    ddata = ddata.loc[ddata['id'] == idnum]
    if isinstance(meal, str):
        ddata = ddata.loc[ddata['meal'] == meal]
        
    ddata['date'] = pd.to_datetime(ddata['date'])
    ddata.set_index('date', inplace = True)
    
    if isinstance(times, tuple):
        ddata = ddata.loc[times[0]:times[1]]
    elif times not in ddata.index:
        ddata = pd.DataFrame()
    else:
        ddata = ddata.loc[times]

    return ddata

def combine(oldser):
    newser = pd.Series()

    for row in oldser.index.drop_duplicates():
            newser.at[row] = np.float64(0)

    for n in newser.index:
        if isinstance(oldser.at[n], int | float | np.float64 | np.int64):
            newser.at[n] = oldser.at[n]
        else:
            cser = oldser.at[n]
            newser.at[n] = cser.sum()
    
    return newser

def combdf(olddf):
    newdf = pd.DataFrame(np.float64(0), columns = olddf.columns, index=olddf.index.drop_duplicates())
    
    for day in newdf.index:
        for col in olddf.columns:
            if isinstance(olddf.loc[day, col], np.float64 | float | int | np.int64):
                newdf.loc[day, col] = olddf.loc[day, col]
            elif len(olddf.loc[day, col]) > 1:
                newdf.loc[day, col] = olddf.loc[day, col].sum()
    
    return newdf

def plotti(parm):
    if parm in ['calories', 'totalfat', 'saturated', 'cholesterol', 'sodium', 'carbohydrates', 'dietaryfiber', 'addedsugar', 'protein']:
        isave = True
        if parm == 'calories':
            plti = 'calorie'
            yl = '[Calories]'
        elif parm in ['cholesterol', 'sodium']:
            yl = '[miligrams]'
            plti = parm
        else:
            yl = '[grams]'
            if parm == 'dietaryfiber':
                plti = 'dietary fiber'
            elif parm == 'addedsugar':
                plti = 'added sugar'
            elif parm == 'totalfat':
                plti = 'total fat'
            elif parm == 'saturated':
                plti = 'saturated fat'
            else:
                plti = parm

    elif parm in ['trans', 'monosaturated', 'polyunsaturated', 'solublefiber', 'sugar']:
        isave = False
        yl = '[grams]'
        if parm in ['trans', 'monounsaturated', 'polyunsaturated']:
            plti = parm+' fat'
        elif parm == 'solublefiber':
            plti = 'soluble fiber'
        else:
            plti = 'sugar'
    elif parm in ['eer', 'bmi']:
        isave = False
        plti = parm.upper()
        if parm == 'eer':
            yl = '[Calories]'
        else:
            yl = '[kg/m^2]'
    else:
        plti = parm.title()
        yl = '[lbs]'
        isave = False
        
    return plti, yl, isave

def color(parameter, ave, rec):
    if parameter in ['cholesterol', 'addedsugar', 'saturatedfat', 'sodium', 'calories']:
        if ave > rec:
            color = 'red'
        else:
            color = 'limegreen'
    elif parameter == 'dietaryfiber':
        if ave > rec:
            color = 'limegreen'
        else:
            color = 'red'
    else:
        color = 'darkorange'
            
    return color

def mp_title(meal, period = tuple):
    if meal is False:
        meal_title = ''
    else:
        meal_title = ' for '+meal.title()

    if period[0] == period[1]:
        period_title = 'Today'
    else:
        period_title = dt.strftime(period[0], '%m/%d/%Y')+' to '+dt.strftime(period[1], '%m/%d/%Y')
    
    return meal_title, period_title

def ingredients_list(flist, fooddf):
    indexes = []
    for m in range(0, len(flist), 2):
        if isinstance(fooddf.at[flist[m], 'ingredients'], str):
            new_flist = fooddf.at[flist[m], 'ingredients'].split('BREAK')
            new_flist.remove('')
            flist += new_flist
            indexes.append(m)
            
    adj = 0

    for n in indexes:
        del flist[n - adj]
        del flist[n - adj]
        adj += 2
    
    return flist

def daily_value(eer):
    multiplier = eer / 2000

    dv_dict = {}

    dv_dict['calories'] = eer
    dv_dict['totalfat'] = multiplier*78 #less than 
    dv_dict['saturated'] = multiplier*20 #less than
    dv_dict['cholesterol'] = 300
    dv_dict['sodium'] = 2300
    dv_dict['carbohydrates'] = multiplier*275
    dv_dict['dietaryfiber'] = 28
    dv_dict['addedsugar'] = multiplier*50
    dv_dict['protein'] = multiplier*50

    return dv_dict

def percentdv(food, dvals):

    percents = {}

    if isinstance(food, pd.Series):
        for n in dvals.keys():
            percents[n] = (food.at[n] / dvals[n])*100
    elif isinstance(food, Day):
        for a, b in food.__dict__.items():
            if a in dvals.keys():
                percents[a] = (b / dvals[a])*100

    return percents

class Profile:
    def __init__(self):
        self.name = ''
        self.birthday = ''
        self.sex = ''
        self.height = float('NaN')
        self.weight = float('NaN')
        self.lifestyle = float('NaN')
        self.bmi = float('NaN')
        self.eer = float('NaN')
        self.goal = float('NaN')
        self.id = float('NaN')
        self.update = ''

    def load_profile(self, data_frame):
        self.name = data_frame['name']
        self.birthday = data_frame['birthday']
        self.sex = data_frame['sex']
        self.height = data_frame['height']
        self.weight = data_frame['weight']
        self.lifestyle = data_frame['lifestyle']
        self.bmi = data_frame['bmi']
        self.eer = data_frame['eer']
        self.goal = data_frame['goal']
        self.id = data_frame['id']
        self.update = data_frame['update']

    def save_data(self, patientdf, file):
        data = []
        for a, b in self.__dict__.items():
            data.append(b)

        try:
            with open(file, 'a') as f:
                writer = csv.writer(f)
                writer.writerow(data)
            print(Style.NORMAL+Fore.GREEN+'\nData saved successfully.')
            print(Fore.BLACK)
        except PermissionError:
            print(Style.NORMAL+Fore.RED+f"\nCould not save new user data because {file!r} is in use. Please close file and try again to save.")
            print(Fore.BLACK)

        patientdf = patientdf._append(self.__dict__, ignore_index = True)

        return patientdf 
    
    def calc_age(self):
        today = datetime.datetime.now()
        diff = today - datetime.datetime.strptime(self.birthday, '%m/%d/%Y') 
        age = diff.days / 365

        return age

    def calc_eer(self):
        meters = self.height*0.0254
        kilos = self.weight*0.453592
        age = self.calc_age()

        if age > 18:
            if self.sex == 'male':
                group = 'men'
            else:
                group = 'women'
        elif age < 18:
            if self.sex == 'male':
                group = 'boys'
            else: 
                group = 'girls'
        else: 
            print(Style.NORMAL+Fore.RED+"\nDue to insufficient data, EER cannot be calculated.")
            print(Fore.BLACK)

        try:
            chart = pd.read_csv('C:/Users/williamsben/OneDrive - Milwaukee School of Engineering/Physical Activity Chart.csv')
        except PermissionError:
            print(Style.NORMAL+Fore.RED+"\nCould not calculate EER because 'Physical Activity Chart.csv' is in use. Please close the file.")
            print(Fore.BLACK)

        pa = chart.loc[self.lifestyle - 1, group]

        match group:
            case 'men':
                self.eer = 662 - 9.53*age + pa*(15.91*kilos + 539.6*meters)
            case 'women':
                self.eer = 354 - 6.91*age + pa*(9.36*kilos + 726*meters)
            case 'boys':
                self.eer = 88.5 - 61.9*age + pa*(26.7*kilos + 903*meters) + 25
            case 'girls':
                self.eer = 135.3 - 30.8*age + pa*(10.0*kilos + 934*meters) + 25
            case _:
                self.eer = float('NaN')

        if self.goal in [1, 3]:
            if self.goal == 1:
                action = 'lose'
            else:
                action = 'gain'

            while True:
                pounds = input(Style.NORMAL+"\nHow many total pounds would you like to "+action+" from your current weight (lbs)?\n\n")
                pounds = split_serv(pounds)

                if pounds[1].isalpha():
                    wgt = pounds[1]
                    pass
                elif math.isnan(pounds[1]):
                    wgt = 'pounds'

                duration = input(Style.NORMAL+"\nOver how much time would you like to "+action+f' {pounds[0]:.0f} {wgt}?\n\n')
                duration = split_serv(duration)

                if duration[1] in ['years', 'year', 'yr', 'yrs', 'y']:
                    weeks = duration[0]*52
                elif duration[1] in ['months', 'month', 'mth', 'mths', 'm']:
                    weeks = (duration[0] / 12)*52
                elif duration[1] in ['weeks', 'week', 'wk', 'wks', 'w']:
                    weeks = duration[0]
                else:
                    print(Style.NORMAL+Fore.RED+"\nInvalid duration entered. Please retry.")
                    print(Fore.BLACK)
                    continue

                pounds_week = pounds[0] / weeks

                if pounds_week > 2:
                    print(Style.NORMAL+Fore.RED+"\nWeight "+action+f" of {pounds[0]:.0f} lbs/{duration[1].strip('s')} is not recommended. Please enter a new plan.")
                    print(Fore.BLACK)
                    continue

                calories = (pounds[0] / weeks)*500

                if self.goal == 1:
                    self.eer -= calories
                elif self.goal == 3:
                    self.eer += calories
                
                break

    def disp_info(self):
        print(Style.NORMAL+Fore.BLACK+"___________"*10)
        print("\nPatient: " + self.name)
        print(f"\nID: {self.id:.0f}")
        print("\nDate of Birth: "+self.birthday)
        print(f"\nAge: {int(self.calc_age())} years old")
        print("\nSex: " + self.sex)
        hgt = conv_height(self.height)
        print("\nHeight: "+str(int(hgt[0]))+"' "+str(int(hgt[1]))+'"')
        print(f"\nWeight: {self.weight:.0f} lbs")
        match self.lifestyle:
            case 1:
                life = 'Sedentary'
            case 2:
                life = 'Low physical activity'
            case 3:
                life = 'Moderate physical activity'
            case 4:
                life = 'High physical activity'
            case _:
                life = 'N/A'
        print("\nLifestyle: "+life)
        print(f"\nBody Mass Index (BMI): {self.bmi:.1f}")
        print(f"\nEstimated Energy Requirement (EER): {self.eer:.1f} Calories")
        match self.goal:
            case 1:
                gl = 'Weight loss'
            case 2:
                gl = 'Weight maintenance'
            case 3:
                gl = 'Weight gain'
            case _:
                gl = 'N/A'
        print("\nGoal: "+gl)
        print("___________"*10)

    def update_profile(self):
        pass

class Food:
    def __init__(self):
        self.name = ''
        self.type = ''
        self.brand = ''
        self.weight = float('NaN')
        self.size = float('NaN')
        self.measurement = ''
        self.calories = float('NaN')
        self.totalfat = float('NaN')
        self.saturated = float('NaN')
        self.trans = float('NaN')
        self.polyunsaturated = float('NaN')
        self.monounsaturated = float('NaN')
        self.cholesterol = float('NaN')
        self.sodium = float('NaN')
        self.carbohydrates = float('NaN')
        self.dietaryfiber = float('NaN')
        self.solublefiber = float('NaN')
        self.sugar = float('NaN')
        self.addedsugar = float('NaN')
        self.protein = float('NaN')
        self.ingredients = ''

    def load_food(self, data_frame):
        self.name = data_frame.loc['name']
        self.type = data_frame.loc['type']
        self.brand = data_frame.loc['brand']
        self.weight = data_frame.loc['weight']
        self.size = data_frame.loc['size']
        self.measurement = data_frame.loc['measurement']
        self.calories = data_frame.loc['calories']
        self.totalfat = data_frame.loc['totalfat']
        self.saturated = data_frame.loc['saturated']
        self.trans = data_frame.loc['trans']
        self.polyunsaturated = data_frame.loc['polyunsaturated']
        self.monounsaturated = data_frame.loc['monounsaturated']
        self.cholesterol = data_frame.loc['cholesterol']
        self.sodium = data_frame.loc['sodium']
        self.carbohydrates = data_frame.loc['carbohydrates']
        self.dietaryfiber = data_frame.loc['dietaryfiber']
        self.solublefiber = data_frame.loc['solublefiber']
        self.sugar = data_frame.loc['sugar']
        self.addedsugar = data_frame.loc['addedsugar']
        self.protein = data_frame.loc['protein']
        self.ingredients = data_frame.loc['ingredients']
    
    def save_data(self, fooddf, file):
        data = []
        for a, b in self.__dict__.items():
            data.append(b)
        
        food_data = load_workbook(file)
        fd_sheet = food_data.active

        row = fd_sheet.max_row+1

        for col, val in enumerate(data, start=1):
            fd_sheet.cell(row=row, column=col, value=val)

        food_data.save(file)
        food_data.close()

        fooddf = fooddf._append(self.__dict__, ignore_index = True)

        return fooddf

    def disp_info(self, eer):
        ser = pd.Series(0)
        for n, m in self.__dict__.items():
            ser[n] = m
        
        values = daily_value(eer)
        dv = percentdv(ser, values)

        print("___________"*10)
        print(Style.BRIGHT+"\nBASIC FACTS")
        print(Style.NORMAL+"\nFood: "+self.name)
        print("\nType: "+self.type)
        print("\nBrand: "+self.brand)
        print("___________"*10)
        print("\nNUTRITION FACTS")
        print("\nServing size\t"+str(self.size)+" "+self.measurement+f" ({self.weight:.0f} g)")
        print(f"\nCalories\t{self.calories:.0f}\t\t\t\t\t{dv['calories']:.0f}%")
        print(f"\nTotal Fat\t{self.totalfat:.0f} g\t\t\t\t\t{dv['totalfat']:.0f}%")
        print(f"\n\tSaturated Fat\t{self.saturated:.0f} g\t\t\t\t{dv['saturated']:.0f}%")
        print(f"\n\tTrans Fat\t{self.trans:.0f} g")
        print(f"\n\tPolyunsaturated Fat\t{self.polyunsaturated:.0f} g")
        print(f"\n\tMonounsaturated Fat\t{self.monounsaturated:.0f} g")
        print(f"\nCholesterol\t{self.cholesterol:.0f} mg\t\t\t\t\t{dv['cholesterol']:.0f}%")
        print(f"\nSodium\t{self.sodium:.0f} mg\t\t\t\t\t\t{dv['sodium']:.0f}%")
        print(f"\nTotal Carbohydrates\t{self.carbohydrates:.0f} g\t\t\t\t{dv['carbohydrates']:.0f}%")
        print(f"\n\tDietary Fiber\t{self.dietaryfiber:.0f} g\t\t\t\t{dv['dietaryfiber']:.0f}%")
        print(f"\n\tSoluble Fiber\t{self.solublefiber:.0f} g")
        print(f"\n\tTotal Sugars\t{self.sugar:.0f} g")
        print(f"\n\t\tIncludes {self.addedsugar:.0f} g Added Sugars\t\t{dv['addedsugar']:.0f}%")
        print(f"\nProtein\t{self.protein:.0f} g\t\t\t\t\t\t{dv['protein']:.0f}%\n")
        print("___________"*10)
    
class Day(Food):
    def __init__(self, date, idnum):
        self.date = date
        self.time = ''
        self.id = idnum
        self.meal = ''
        super().__init__()
        del self.name
        del self.type
        del self.brand
        del self.weight
        del self.size
        del self.measurement 
        del self.ingredients
        self.foods = ''

    def load_date(self, data_frame):
        self.time = data_frame.loc['time']
        self.meal = data_frame.loc['meal']
        self.calories = data_frame.loc['calories']
        self.totalfat = data_frame.loc['totalfat']
        self.saturated = data_frame.loc['saturated']
        self.trans = data_frame.loc['trans']
        self.polyunsaturated = data_frame.loc['polyunsaturated']
        self.monounsaturated = data_frame.loc['monounsaturated']
        self.cholesterol = data_frame.loc['cholesterol']
        self.sodium = data_frame.loc['sodium']
        self.carbohydrates = data_frame.loc['carbohydrates']
        self.dietaryfiber = data_frame.loc['dietaryfiber']
        self.solublefiber = data_frame.loc['solublefiber']
        self.sugar = data_frame.loc['sugar']
        self.addedsugar = data_frame.loc['addedsugar']
        self.protein = data_frame.loc['protein']
        self.foods = data_frame.loc['foods']

    def save_data(self, daydf, file):
        return super().save_data(daydf, file)

    def summary(self, eer, inputdf, place=str):
        if place in ['anl', 'ldg']:
            today = datetime.date.today()
            today = today.strftime('%Y-%m-%d')

            ddata = date_index(self.id, today, inputdf, False)
        
            if ddata.empty:
                print(Style.NORMAL+"\n\nDAILY SUMMARY\n\nDate: "+self.date+"\n\nLast meal entry: none\n\nMeals eaten: none\n\nTotal calories consumed: 0 Calories")
            else:
                dframe = pd.Series()

                if isinstance(ddata, pd.DataFrame):
                    for n in ddata.columns:
                        if n == 'time':
                            dframe[n] = ddata.iloc[-1, 0]
                        else:
                            dframe[n] = ddata[n].sum()
                else:
                    dframe = ddata

                self.load_date(dframe)

                print(Style.NORMAL+"\nDAILY SUMMARY")
                print("\nDate: "+self.date)
                print("\nLast meal entry: "+self.time)

                meals = ''

                for n in ['breakfast', 'lunch', 'dinner', 'snack', 'dessert']:
                    if n in self.meal:
                        meals += n+', '
                meals = meals.removesuffix(', ')

                print("\nMeals eaten: "+meals)

                if place == 'ldg':
                    print(f"\nTotal calories consumed: {self.calories:.0f} Calories")

        if place in ['anl', 'meal']:
            dailyvalues = daily_value(eer)
            percents = percentdv(self, dailyvalues)
            if place == 'meal':
                print("___________"*10)
                print(Style.BRIGHT+"\nMEAL SUMMARY")

            print(Style.NORMAL+"\nFoods consumed:")

            meals = self.foods.split('BREAK')

            for n in range(0, len(meals) - 1, 2):
                print("\n-"+meals[n+1]+" of "+meals[n])

            print("\n\t\t\t\t\t\t"+"* * "*5)
            print(f"\nTotal calories: {self.calories:.0f} out of {eer:.0f} Calories ({percents['calories']:.0f}% DV)")
            print(f"\nTotal fat: {self.totalfat:.0f} g out of {dailyvalues['totalfat']:.0f} g ({percents['totalfat']:.0f}% DV)")
            print(f"\nSaturated fat: {self.saturated:.0f} g out of {dailyvalues['saturated']:.0f} g ({percents['saturated']:.0f}% DV)")
            print(f"\nTrans fat: {self.trans:.0f} g")
            print(f"\nPolyunsaturated fat: {self.polyunsaturated:.0f} g")
            print(f"\nMonounsaturated fat: {self.monounsaturated:.0f} g")
            print(f"\nCholesterol: {self.cholesterol:.0f} mg out of {dailyvalues['cholesterol']:.0f} mg ({percents['cholesterol']:.0f}% DV)")
            print(f"\nSodium: {self.sodium:.0f} mg out of {dailyvalues['sodium']:.0f} mg ({percents['sodium']:.0f}% DV)")
            print(f"\nCarbohydrates: {self.carbohydrates:.0f} g out of {dailyvalues['carbohydrates']:.0f} g ({percents['carbohydrates']:.0f}% DV)")
            print(f"\nDietary fiber: {self.dietaryfiber:.0f} g out of {dailyvalues['dietaryfiber']:.0f} g ({percents['dietaryfiber']:.0f}% DV)")
            print(f"\nSoluble fiber: {self.solublefiber:.0f} g")
            print(f"\nTotal sugar: {self.sugar:.0f} g")
            print(f"\nAdded sugar: {self.addedsugar:.0f} g out of {dailyvalues['addedsugar']:.0f} g ({percents['addedsugar']:.0f}% DV)")
            print(f"\nProtein: {self.protein:.0f} g out of {dailyvalues['protein']:.0f} g ({percents['protein']:.0f}% DV)")
            print("___________"*10)

def adjust(aclass, food_string, dataf):
        food_list = food_string.split('BREAK')
        food_list.remove('')

        adjusted = pd.DataFrame(0, index = [0], columns = vars(aclass).keys())

        fooddf = dataf.copy()
        fooddf.drop_duplicates(subset = 'name', inplace = True)
        fooddf.set_index('name', inplace = True)
        fooddf.replace(float('NaN'), 0, inplace = True)
    
        for n in range(0, len(food_list)-1, 2):
            servings = split_serv(food_list[n+1])
            multi = servings[0] / fooddf.at[food_list[n], 'size']

            for col in adjusted.columns:
                if col in fooddf.columns:
                    val = fooddf.loc[food_list[n], col]
                    if (isinstance(val, float | int | np.int64 | np.float64)): 
                        adjusted[col] += val*multi

                    elif isinstance(val, str):
                        if val.isnumeric():
                            if '.' in val:
                                adjusted[col] += float(val)*multi
                            elif '/' in val:
                                vals = val.split('/')
                                adjusted[col] += (val[0] / vals[1])*multi
                            else:
                                adjusted[col] += (int(val))*multi
                        else:
                            adjusted[col] = val

        if isinstance(aclass, Day):
            adjusted['time'] = aclass.time
            adjusted['meal'] = aclass.meal
            adjusted['foods'] = food_string
            aclass.load_date(adjusted.loc[0])
        elif isinstance(aclass, Food):
            adjusted['name'] = aclass.name
            adjusted['type'] = 'other'
            adjusted['brand'] = 'none'
            adjusted['size'] = aclass.size
            adjusted['measurement'] = aclass.measurement
            adjusted['ingredients'] = food_string
            aclass.load_food(adjusted.loc[0])

        return adjusted
